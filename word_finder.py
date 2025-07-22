import os
import PyPDF2
from docx import Document
from striprtf.striprtf import rtf_to_text
#import PyInstaller
import openpyxl
#from tkinter_gui import tk
import tkinter as tk
from tkinter import ttk #para barra de carga





class finder():
    def __init__(self, output_text_1, output_text_2):
        self.output_text = output_text_1
        self.output_text_2 = output_text_2
        self.file_check= {}

    def buscar_en_txt(self, ruta, palabra):
        self.file_check[ruta] = False
        found = False
        
        with open(ruta, 'r', encoding='utf-8', errors='ignore') as f:
            for i, linea in enumerate(f, 1):
                if palabra.lower() in linea.lower():
                    print(f"[TXT] {ruta} (línea {i}): {linea.strip()}")
                    self.output_text.insert(tk.END, f"Palabra")
                    self.output_text.insert(tk.END, f" {palabra}", "negrita")
                    self.output_text.insert(tk.END, f" encontrada\n")
                    self.output_text.insert(tk.END, f"[TXT] {ruta} (línea {i}): {linea.strip()}\n\n")
                    self.file_check[ruta] = True
                    found = True
        if found == False:
            print(f"[TXT] {ruta} (línea {i}): {linea.strip()}")
            self.output_text_2.insert(tk.END, f"Palabra")
            self.output_text_2.insert(tk.END, f" {palabra}", "negrita")
            self.output_text_2.insert(tk.END, f" NO encontrada\n")
            self.output_text_2.insert(tk.END, f"[TXT] {ruta} (línea {i}): {linea.strip()}\n\n")

    def buscar_en_pdf(self, ruta, palabra):
        self.file_check[ruta] = False
        try:
            found = False
            with open(ruta, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                for i, page in enumerate(reader.pages):
                    texto = page.extract_text() or ""
                    if palabra.lower() in texto.lower():
                        print(f"[PDF] {ruta} (página {i+1}): contiene la palabra")
                        self.output_text.insert(tk.END, f"Palabra")
                        self.output_text.insert(tk.END, f" {palabra}", "negrita")
                        self.output_text.insert(tk.END, f" encontrada\n")
                        self.output_text.insert(tk.END, f"[PDF] {ruta} (página {i+1})\n\n")
                        self.file_check[ruta] = True
                        found = True
            if found == False:
                print(f"[PDF] {ruta}. Palabra no encontrada")
                self.output_text_2.insert(tk.END, f"Palabra")
                self.output_text_2.insert(tk.END, f" {palabra}", "negrita")
                self.output_text_2.insert(tk.END, f" NO encontrada\n")
                self.output_text_2.insert(tk.END, f"[PDF] {ruta}\n\n")

        except:
            print(f"[PDF] No se pudo leer {ruta}")
            self.output_text_2.insert(tk.END, f"[PDF] No se pudo leer {ruta}\n\n")

    def buscar_en_docx(self, ruta, palabra):
        self.file_check[ruta] = False
        found = False
        try:
            doc = Document(ruta)
            for i, p in enumerate(doc.paragraphs):
                if palabra.lower() in p.text.lower():
                    print(f"[DOCX] {ruta} (párrafo {i+1}): {p.text.strip()}")
                    self.output_text.insert(tk.END, f"\nPalabra")
                    self.output_text.insert(tk.END, f" {palabra}", "negrita")
                    self.output_text.insert(tk.END, f" encontrada\n")
                    self.output_text.insert(tk.END, f"[DOCX] {ruta} (párrafo {i+1}): {p.text.strip()}\n\n")
                    self.file_check[ruta] = True
                    found = True
            
            if found == False:
                print(f"[DOCX] {ruta}. Palabra no encontrada")
                self.output_text_2.insert(tk.END, f"Palabra")
                self.output_text_2.insert(tk.END, f" {palabra}", "negrita")
                self.output_text_2.insert(tk.END, f" NO encontrada\n")
                self.output_text_2.insert(tk.END, f"[DOCX] {ruta}\n\n")
        except:
            print(f"[DOCX] No se pudo leer {ruta}")
            self.output_text_2.insert(tk.END, f"[DOCX] No se pudo leer {ruta}")

    def buscar_en_xlsx(self, ruta, palabra):
        self.file_check[ruta] = False
        found = False
        try:
            wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
            for hoja in wb.sheetnames:
                ws = wb[hoja]
                for fila in ws.iter_rows(values_only=True):
                    for celda in fila:
                        if isinstance(celda, str) and palabra.lower() in celda.lower():
                            print(f"[XLSX] {ruta} (hoja '{hoja}'): contiene la palabra")
                            self.output_text.insert(tk.END, f"\nPalabra")
                            self.output_text.insert(tk.END, f" {palabra}", "negrita")
                            self.output_text.insert(tk.END, f" encontrada\n")
                            self.output_text.insert(tk.END, f"[XLSX] {ruta} (hoja '{hoja}')\n\n")
                            self.file_check[ruta] = True
                            found = True
                            break
            if found == False:
                print(f"[XLSX] {ruta}. Palabra NO encontrada")
                self.output_text_2.insert(tk.END, f"Palabra")
                self.output_text_2.insert(tk.END, f" {palabra}", "negrita")
                self.output_text_2.insert(tk.END, f" NO encontrada\n")
                self.output_text_2.insert(tk.END, f"[XLSX] {ruta}.\n\n")
            
        except:
            print(f"[XLSX] No se pudo leer {ruta}")
            self.output_text_2.insert(tk.END, f"[XLSX] No se pudo leer {ruta}")

    def buscar_en_rtf(self, ruta, palabra):
        self.file_check[ruta] = False
        with open(ruta, "r", encoding="utf-8", errors="ignore") as f:
            contenido_rtf = f.read()
            texto = rtf_to_text(contenido_rtf)
            if palabra.lower() in texto.lower():
                print(f"[RTF] {ruta}  contiene la palabra")
                self.output_text.insert(tk.END, f"\nPalabra")
                self.output_text.insert(tk.END, f" {palabra}", "negrita")
                self.output_text.insert(tk.END, f" encontrada\n")
                self.output_text.insert(tk.END, f"[RTF] {ruta}\n\n")
                self.file_check[ruta] = True
                return
            
            print(f"[RTF] {ruta}. Palabra NO encontrada")
            self.output_text_2.insert(tk.END, f"Palabra")
            self.output_text_2.insert(tk.END, f" {palabra}", "negrita")
            self.output_text_2.insert(tk.END, f" NO encontrada\n")
            self.output_text_2.insert(tk.END, f"[RTF] {ruta}.\n\n")  

            



    def find_word_in_file(self, directorio, cadena):
        
        for root, _,files in os.walk(directorio):
            for file in files:
                ruta = os.path.join(root, file)
                #self.file_check[ruta] = False
                ext = file.lower().split('.')[-1]

                if ext == 'txt':
                    self.buscar_en_txt(ruta, cadena)
                elif ext == 'pdf':
                    self.buscar_en_pdf(ruta, cadena)
                elif ext == 'docx':
                    self.buscar_en_docx(ruta, cadena)
                elif ext == 'xlsx':
                    self.buscar_en_xlsx(ruta, cadena)
                elif ext == "rtf":
                    self.buscar_en_rtf(ruta, cadena)    
                
                else:
                    print(f"Formato no soportado de {ruta}")

